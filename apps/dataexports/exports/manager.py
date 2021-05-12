from cc_courses.models import Activity, Organizer
from django.http import HttpResponseNotFound, HttpResponse
from openpyxl import Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill
from django.db.models import Q
import json


class ExportManager:
    def __init__(self):
        self.ignore_errors = False
        self.subsidy_period = None
        self.subsidy_period_range = None
        self.error_message = set()

    def return_404(self, message=""):
        """When the exported data has to fit a specific format, there
        are many cases in which we need to stop the generation and tell
        the user that something needs to be fixed.
        This will show a blank page with the message.
        """
        if message:
            self.error_message.add(message)
        message = "<h1>Error al generar el document</h1>" + " ".join(
            self.error_message)
        return HttpResponseNotFound(message)

    def import_correlations(self, file_path):
        try:
            file_object = open(file_path, 'r')
            self.correlations = json.load(file_object)
        except FileNotFoundError:
            print(file_path + " not found. ")

    def get_correlation(self, correlated_field, original_data, subsidy_period=2019):
        """When exporting data, we might need to make the exported data
         fit specific requirements. For example, we store the field
         'axis' as 'A', 'B', but the strings we actually need to
         show are:
         'A) Diagnosi i visibilització', 'B) Creació i desenvolupament'

         We have these correlations in a json file and loaded at self.correlations.

        This function is a wrapper to get those.
        """
        try:
            new_data = self.correlations[correlated_field][original_data]
        except KeyError:
            self.error_message.add(
                "<p>El document no s'ha pogut generar perquè s'ha intentat aplicar aquesta correlació:</p"
                "<ul><li>Convocatòria: {}</li><li>Camp: {}</li><li>Dada original: {}</li></ul>"
                "<p>Però no s'ha trobat.</p>".format(subsidy_period, correlated_field, original_data))
            return None
        return new_data


class ExcelExportManager(ExportManager):
    def __init__(self):
        super().__init__()
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

        self.stages_obj = None

        self.row_number = 1
        self.number_of_activities = 0
        self.number_of_stages = 0
        self.number_of_nouniversitaris = 0
        self.number_of_founded_projects = 0

        self.correlations = dict()
        self.organizers = dict()  # Camp per Ateneu / Cercle
        self.d_organizer = None

        # La majoria d'ateneus volen que hi hagi una sola actuació per un
        # projecte encara que hagi tingut diferents tipus d'acompanyament.
        # CoopCamp (i potser algun altre?) volen separar-ho per itineraris,
        # de manera que hi hagi una actuació per l'itinerari de Nova Creació i
        # una pel de Consolidació.
        # Per defecte ho definim per 1 itinerari.
        self.stages_groups = {
            1: 'nova_creacio',
            2: 'nova_creacio',
            6: 'nova_creacio',
            7: 'nova_creacio',
            8: 'nova_creacio',
            9: 'nova_creacio'  # Era Incubació
        }

    def return_document(self, name):
        """ Attention: non-ascii characters in the name will cause
        an encoding error with gunicorn.
        Haven't tried it with a proxy under apache, in theory should
        work."""
        if len(self.error_message) > 0 and self.ignore_errors is False:
            return self.return_404()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'), name=name,
        )
        self.workbook.save(response)
        return response

    def get_sessions_obj(self, for_minors=False):
        return Activity.objects.filter(
            Q(date_start__range=self.subsidy_period.range, for_minors=for_minors) &
            (
                Q(cofunded__isnull=True) | (
                    Q(cofunded__isnull=False) & Q(cofunded_ateneu=True)
                )
            )
        )

    def create_columns(self, columns):
        """ create_columns

        Expects an iterable containing tuples with the name and the
        width of each column, like this:
        columns = [
            ("First", 40),
            ("Second", 70),
        ]
        """
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            column_letter = get_column_letter(col_num)
            column_dimensions = self.worksheet.column_dimensions[column_letter]
            column_dimensions.font = Font(name="ttf-opensans", size=9)
            column_dimensions.width = column_width
            cell.font = Font(bold=True, name="ttf-opensans", size=9)
            cell.border = Border(bottom=Side(border_style="thin", color='000000'))
            cell.value = str(column_title)

    def fill_row_data(self, row):
        """ fill_row_data

        Populates the columns of a given row with each of the values.
        Expects an iterable with the data for each row:
        row = [
            "first value",
            "second value",
        ]

        Optionally, values can be a tuple to mark the cell as error.
        That will fill the cell with red.
        row = [
            "first value",
            ("second value", True),
        ]
        """
        for col_num, cell_value in enumerate(row, 1):
            cell = self.worksheet.cell(row=self.row_number, column=col_num)
            if isinstance(cell_value, tuple):
                error_mark = cell_value[1]
                if error_mark:
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                cell_value = cell_value[0]
            cell.value = cell_value if isinstance(cell_value, int) else str(cell_value)

    def import_organizers(self):
        # Not forcing any order because we want it in the same order that
        # they see (which should be by ID)
        orgs = Organizer.objects.all()
        if not orgs:
            self.organizers.update({
                0: 'Ateneu'
            })
        else:
            i = 0
            for org in orgs:
                if i == 0:
                    cercle = 'Ateneu'
                else:
                    cercle = f"Cercle {i}"
                self.organizers.update({
                    org.id: cercle
                })
                i += 1
        self.d_organizer = list(self.organizers.keys())[0]

    def get_organizer(self, organizer):
        if not organizer:
            return self.organizers[self.d_organizer]
        return self.organizers[organizer.id]
